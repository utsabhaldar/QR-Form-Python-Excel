from flask import Flask, render_template, request, redirect, send_file, url_for
# from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import qrcode
import os

app = Flask(__name__)
# CORS(app)

# Path to the Excel file
excel_file = "user_data.xlsx"

# Ensure Excel file exists
def ensure_excel_exists():
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Sex", "Age", "Contact", "Email"])
        workbook.save(excel_file)

# Save data to Excel
def save_to_excel(data):
    ensure_excel_exists()
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    sheet.append(data)
    workbook.save(excel_file)

# Home route for the form
@app.route('/', methods=['GET', 'POST'])
def form():
    if request.method == 'POST':
        name = request.form['name']
        sex = request.form['sex']
        age = request.form['age']
        contact = request.form['contact']
        email = request.form['email']
        save_to_excel([name, sex, age, contact, email])
        return redirect(url_for('thank_you'))
    return render_template('form.html')

@app.route('/thank-you')
def thank_you():
    return "<h2>Thank you for submitting the form!</h2>"

# Admin route to download Excel file
@app.route('/admin/download')
def download_excel():
    return send_file(excel_file, as_attachment=True)

# Admin route to generate QR code
@app.route('/admin/generate-qr')
def generate_qr():
    url = url_for('form', _external=True)
    qr_img = qrcode.make(url)
    qr_img_path = "static/qr_code.png"
    qr_img.save(qr_img_path)
    return f'<img src="/{qr_img_path}" alt="QR Code"> <br> <a href="/{qr_img_path}" download>Download QR Code</a>'

if __name__ == '__main__':
    app.run(debug=True)
